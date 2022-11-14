import pandas as pd
import openpyxl
import pymorphy2
from pullenti_wrapper.langs import (
    set_langs,
    RU
)
set_langs([RU])
from pullenti_wrapper.processor import (
    Processor,
    GEO,
    ADDRESS
)
from pullenti_wrapper.referent import Referent

addr = []

def display_shortcuts(referent, level=0):
    tmp = {}
    a = ""
    b = ""
    for key in referent.__shortcuts__:
        value = getattr(referent, key)
        if value in (None, 0, -1):
            continue
        if isinstance(value, Referent):
            display_shortcuts(value, level + 1)
        else:
            if key == 'type':
                a = value
            if key == 'name':
                b = value
                # print('ok', value)
            if key == 'house':
                a = "дом"
                b = value
                tmp[a] = b
            if key == 'flat':
                a = "квартира"
                b = value
                # print('ok', value)
                tmp[a] = b
            if key == 'corpus':
                a = "корпус"
                b = value
                tmp[a] = b
    tmp[a] = b
    addr.append(tmp)

def exProcess(filename):
    processor = Processor([GEO, ADDRESS])
    file = "tables/" + filename
    xl = openpyxl.load_workbook(file, data_only=True)
    sheets = xl.sheetnames
    print(sheets)
    res = list()
    respos = 1
    if len(sheets) < 7:
        amount = len(sheets)
    else:
        amount = len(sheets) - 7
    usedWorks = list()
    wordslist = list()
    for sheetnum in range(amount):
        sheet = xl[sheets[sheetnum]]
        singlecostcolumn = -1
        allcostcolumn = -1
        measuringunitcolumn = -1
        numberofunitscolumn = -1
        codecol = -1
        address = " "
        code = 0
        singlecost = 0
        allcost = 0
        measure = "m"
        unit = 0
        whitelist = [
            "Замена",
            "Установка",
            "Демонтаж",
            "Изготовление",
            "Устройство",
            "Ремонт",
            "Исправление",
            "Размещение",
            "Посев",
            "Нанесение",
            "Монтаж",
            "Окрашивание",
            "Исправление",
            "Окраска",
            "Разработка"
        ]
        banList = [
            "Масса",
            "Стоимость",
            "Сплошная",
            "Щебень",
            "Смеси",
            "Механизированная",
            "Доски",
            "Шифр",
            "Локальная",
            "Плиты",
            "Пигменты",
            "Итого",
            "Всего"
        ]




        costex = 0

        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):

                cellval = sheet.cell(row=row, column=column).value
                if type(cellval) is not None and type(cellval) is str:
                    result = processor(cellval)
                    if len(result.matches) != 0:
                        referent = result.matches[0].referent
                        display_shortcuts(referent)
                        print(addr)
                        for elem in addr:
                            if 'улица' in elem.keys():
                                address = elem['улица']
                                for elem in addr:
                                    if 'дом' in elem.keys():
                                        address += " " + elem['дом']
                    addr.clear()
                    if codecol == -1 and "шифр" in cellval.lower():
                        codecol = column
                    if measuringunitcolumn == -1 and ("единица изм" in cellval.lower() or measuringunitcolumn == -1 and "ед. изм" in cellval.lower()):
                        measuringunitcolumn = column
                    if numberofunitscolumn == -1 and ("кол-во" in cellval.lower() or "количество" in cellval.lower()):
                        numberofunitscolumn = column
                    if costex == 1 and 'руб.' in cellval.lower() and allcostcolumn == -1:
                        allcostcolumn = column
                        costex = 2
                    if costex == 0 and 'руб.' in cellval.lower() and singlecostcolumn == -1:
                        singlecostcolumn = column
                        costex = 1
                if cellval is not None and\
                        type(cellval) != int and\
                        type(cellval) != float:
                    currentval = cellval.lower().split()
                    if "раздел:" in currentval or "раздел" in currentval:
                        print(usedWorks)
                        usedWorks.clear()
                    if len(cellval.split()) > 4 and \
                            len(cellval) > 15 and \
                            (column != 1 and \
                                cellval not in wordslist and\
                            type(sheet.cell(row=row, column=column - 1).value) is not None and \
                             cellval.split()[0] not in usedWorks and \
                             cellval.split()[0] in whitelist and \
                             not cellval[0].isdigit()):
                                if measuringunitcolumn != -1:
                                    measure = sheet.cell(row=row, column=measuringunitcolumn).value
                                if numberofunitscolumn != -1:
                                    unit = sheet.cell(row=row, column=numberofunitscolumn).value
                                if codecol != -1:
                                    code = sheet.cell(row=row, column=codecol).value
                                currow = row
                                if singlecostcolumn != -1:
                                    while type(sheet.cell(row=currow, column=singlecostcolumn).value) is not float:
                                        currow += 1
                                        if currow - row > 15:
                                            break
                                    singlecost = sheet.cell(row=currow, column=singlecostcolumn).value
                                currow = row
                                if allcostcolumn != -1:
                                    while type(sheet.cell(row=currow, column=allcostcolumn).value) is not float:
                                        currow += 1
                                        if currow - row > 15:
                                            break
                                    allcost = sheet.cell(row=currow, column=allcostcolumn).value
                                usedrow = row
                                focalWord = cellval.split()[0]
                                usedWorks.append(focalWord)
                                wordslist.append(cellval)
                                word = cellval
                                dct = dict()
                                if 'Замена' == focalWord or\
                                    'Устройство' == focalWord or\
                                    'Установка' == focalWord or\
                                    'Ремонт' == focalWord or\
                                    'Исправление' == focalWord:
                                    dct.update({"kpgz": "02.06.05.01 РАБОТЫ РЕМОНТНО-ВОССТАНОВИТЕЛЬНЫЕ СВЯЗАННЫЕ С ЭЛЕМЕНТАМИ БЛАГОУСТРОЙСТВА ТЕРРИТОРИЙ"})
                                if 'Изготовление' == focalWord:
                                    dct.update({"kpgz": "02.03.03.10 ОБУСТРОЙСТВО ТЕРРИТОРИЙ ЭЛЕМЕНТАМИ ИНЖЕНЕРНЫХ КОММУНИКАЦИЙ"})
                                if 'Демонтаж' == focalWord:
                                    dct.update({"kpgz": "02.03.03.06 ОБУСТРОЙСТВО МАФ ТЕРРИТОРИЙ"})
                                if 'Размещение' == focalWord or 'Разработка' == focalWord:
                                    dct.update({"kpgz": "02.03.04 СТРОИТЕЛЬСТВО ПАРКОВ, МЕСТ ОТДЫХА И ДОСУГА"})
                                if 'Посев' == focalWord:
                                    dct.update({"kpgz": "02.03.03.03 ОЗЕЛЕНЕНИЕ ТЕРРИТОРИЙ"})
                                if 'Нанесение' == focalWord:
                                    dct.update({"kpgz": "02.06.05.02 РАБОТЫ РЕМОНТНО-ВОССТАНОВИТЕЛЬНЫЕ СВЯЗАННЫЕ С ЭЛЕМЕНТАМИ СРЕДСТВ ОБЕСПЕЧЕНИЯ ТРАНСПОРТНОЙ БЕЗОПАСНОСТИ ТЕРРИТОРИЙ"})
                                if 'Монтаж' == focalWord:
                                    dct.update({"kpgz": "02.03.03.07 ОБУСТРОЙСТВО ТЕРРИТОРИЙ СРЕДСТВАМИ ОБЕСПЕЧЕНИЯ ТРАНСПОРТНОЙ БЕЗОПАСНОСТИ"})
                                if 'Окрашивание' == focalWord or 'Окраска' == focalWord:
                                    dct.update({"kpgz": "02.03 РАБОТЫ ПО БЛАГОУСТРОЙСТВУ ТЕРРИТОРИЙ"})
                                if 'Исправление' == focalWord:
                                    dct.update({"kpgz": "02.03.03.08 ОБУСТРОЙСТВО ТЕРРИТОРИЙ ЭЛЕМЕНТАМИ ОРГАНИЗАЦИИ РЕЛЬЕФА"})
                                dct.update({"num": respos})
                                dct.update({"addr": address})
                                dct.update({"id": usedrow})
                                if code != 0:
                                    dct.update({"code": code})
                                else:
                                    dct.update({"code": respos})
                                dct.update({"spgz": focalWord})
                                dct.update({"name": word})
                                dct.update({"singlecost": singlecost})
                                dct.update({"allcost": allcost})
                                dct.update({"measure": measure})
                                dct.update({"unit": unit})
                                res.append(dct.copy())
                                respos += 1
        print(usedWorks)
    return res

